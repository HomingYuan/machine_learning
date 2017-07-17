#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import Workbook,load_workbook
import pandas as pd
from openpyxl.styles import Color, Fill
import numpy as np
from openpyxl.styles import Font, Color
from openpyxl.styles import colors

# wb = Workbook()
wb = load_workbook('km.xlsx') # 导入excel表
ws = wb.active

# print(wb.get_sheet_names())
"""
for sheet in wb.get_sheet_names():
    print(sheet.title())
"""

# print(ws.max_row) # 最大行号
# print(ws.max_column) # 最大列号
# print(ws[ws.max_row][0].value)
max_row = "A" +str(ws.max_row)
new_row = "B" +str(ws.max_row+1)
df = pd.read_excel('km.xlsx', sheetname='Sheet1')
l2 = df['x2'].values.T
ws[new_row] = np.array(l2).mean()
# ws['B10000'] = 1
# print(np.array(l2).mean())
mean_value = np.array(l2).mean()
for i in range(2,ws.max_row+1):
    cells_value = ws["B" + str(i)].value
    if cells_value <  mean_value:
        ws['B'+str(i)].font =  Font(color=colors.GREEN)
    else:
        ws['B' + str(i)].font = Font(color=colors.RED)
# ws.cell(ws.max_row+1,1).value = np.array(ws['A2':max_row]).mean()
wb.save('test.xlsx')