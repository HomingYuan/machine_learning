#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


wb = Workbook()
ws = wb.active
df = pd.read_excel('km.xlsx', sheetname='Sheet1')
for r in dataframe_to_rows(df, index=True, header=True):
    ws.append(r)
"""
for cell in ws['A'] + ws[1]:
    cell.style = 'Pandas'
"""
wb.save("pandas_openpyxl.xlsx")