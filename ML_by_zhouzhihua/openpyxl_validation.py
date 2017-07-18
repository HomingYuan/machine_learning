#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
ws = wb.active

dv = DataValidation(type="list", formula1='"Dog,Cat,Bat"', allow_blank=True)
dv.error ='Your entry is not in the list'
dv.errorTitle = 'Invalid Entry'
dv.prompt = 'Please select from the list'
dv.promptTitle = 'List Selection'
ws.add_data_validation(dv)
c1 = ws["A1"]
c1.value = "Dog"
dv.add(c1)
c2 = ws["A2"]
c2.value = "An invalid value"
dv.add(c2)
dv.ranges.append('B1:B1048576')
wb.save("test3.xlsx")