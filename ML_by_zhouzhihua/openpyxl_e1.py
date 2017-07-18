#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.formatting.rule import Rule,DataBar, FormatObject


wb = Workbook()
ws = wb.active

redFill = PatternFill(start_color='EE1111',
                        end_color='EE1111',fill_type='solid')

ws.conditional_formatting.add('A1:A10',ColorScaleRule(start_type='min', start_color='AA0000',
                                end_type='max', end_color='00AA00'))

ws.conditional_formatting.add('B1:B10',ColorScaleRule(start_type='percentile', start_value=10, start_color='AA0000',
                                mid_type='percentile', mid_value=50, mid_color='0000AA',
                                end_type='percentile', end_value=90, end_color='00AA00'))

ws.conditional_formatting.add('C2:C10',CellIsRule(operator='lessThan', formula=['C$1'], stopIfTrue=True, fill=redFill))

ws.conditional_formatting.add('D2:D10',CellIsRule(operator='between', formula=['1','5'], stopIfTrue=True, fill=redFill))

ws.conditional_formatting.add('E1:E10',FormulaRule(formula=['ISBLANK(E1)'], stopIfTrue=True, fill=redFill))

myFont = Font()

myBorder = Border()

ws.conditional_formatting.add('E1:E10',FormulaRule(formula=['E1=0'], font=myFont, border=myBorder, fill=redFill))

red_text = Font(color="9C0006")

red_fill = PatternFill(bgColor="FFC7CE")

dxf = DifferentialStyle(font=red_text, fill=red_fill)

rule = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)

rule.formula = ['NOT(ISERROR(SEARCH("highlight",A1)))']

ws.conditional_formatting.add('A1:F40', rule)

wb.save('test2.xlsx')