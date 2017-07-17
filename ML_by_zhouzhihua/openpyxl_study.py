#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
@author: Homing
@software: PyCharm Community Edition
@file: openpyxl_study.py
@time: 2017/7/17 21:27
"""
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from openpyxl import Workbook
"""
font = Font(name='Calibri',size=11,
            bold=False, italic=False,
            vertAlign=None,underline='none',
            strike=False,color='FF00000')

fill = PatternFill(fill_type=None,start_color='FFFFFFFF', end_color='FF000000')

border = Border(left=Side(border_style=None,
color='FF000000'),
right=Side(border_style=None,
color='FF000000'),
top=Side(border_style=None,
color='FF000000'),
bottom=Side(border_style=None,
color='FF000000'),
diagonal=Side(border_style=None,
color='FF000000'),
diagonal_direction=0,
outline=Side(border_style=None,
color='FF000000'),
vertical=Side(border_style=None,
color='FF000000'),
horizontal=Side(border_style=None,
color='FF000000'))

alignment=Alignment(horizontal='general',
vertical='bottom',
text_rotation=0,
wrap_text=False,
shrink_to_fit=False,
indent=0)
number_format = 'General'

protection = Protection(locked=True,
hidden=False)

"""
wb = Workbook()
ws = wb.active
a1 = ws['A1']

d4 = ws['D4']
ft = Font(color=colors.RED)
a1.font = ft
d4.font = ft
ws['A1'] = 'RED'
ws['D4'] = 'RED'
# a1.font.italic = True # is not allowed
a1.font = Font(color=colors.RED, italic=True) # the change only affects A1
ft1 = Font(name='Arial', size=14,bold=True,underline="single",strike=True,color='FF0000')
ft2 = ft1
b2 = ws['B2']
c3 = ws['C3']

b2.font = ft1
c3.font =ft2
ws['b2'] = 'My style'
ws['c3'] = 'My style1'
wb.save('test1.xlsx')


